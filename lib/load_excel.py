from openpyxl import load_workbook


def load_table(filename, head_row = 0):
    """
    Загрузка таблиц (с каждого листа) из файла excel.
    :param filename: путь к файлу с таблицам.
    :param head_row: номер строки с заголовком.
    :return: загруженные таблицы.
    """
    wb = load_workbook(filename=filename, read_only=True, data_only=True)

    data_wb = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        rows = tuple(ws.rows)
        head = rows[head_row]
        table = rows[(head_row + 1):]

        data_ws = []
        for row in table:
            data_row = {}
            for i in range(len(head)):
                data_row[head[i].value] = row[i].value

            flag_add = False
            for value in data_row.values():
                if value != None:
                    flag_add = True
                    break

            if flag_add: data_ws.append(data_row)

        data_wb[sheetname] = data_ws

    return data_wb

def load_sheets_name(filename):
    wb = load_workbook(filename=filename, read_only=True, data_only=True)
    return wb.sheetnames
