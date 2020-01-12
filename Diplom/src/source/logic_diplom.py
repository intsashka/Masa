from openpyxl import load_workbook

import lib.load_excel as load_excel
import lib.date_to_rus_string as date_to_rus_string
import lib.make_pdf as make_pdf

from settings import PATH_TO_DB_SCHOOL_SUBJECTS


def check_head(row):
    try:
        row = list(map(lambda it: it.value, row))

        row.index('ФИО'),
        row.index('Муниципалитет'),
        row.index('Учебное заведение'),
        row.index('Учебное заведение (полностью)'),
        row.index('Класс'),
        row.index('Тип диплома')

        return True
    except:
        return False


def find_head(rows):
    i = 0
    while not check_head(rows[i]):
        i += 1
        if i > 1000:
            raise Exception('#1')

    return i


def load_peoples(filename, sheetnames):
    wb = load_workbook(filename=filename, read_only=True, data_only=True)

    sheetnames = sheetnames.split('$')
    sheetnames = list(map(lambda it: it.strip(), sheetnames))
    sheetnames = list(filter(lambda it: it != "", sheetnames))

    peoples = []
    for sheetname in sheetnames:
        ws = wb[sheetname]
        rows = tuple(ws.rows)

        i = find_head(rows)

        table = load_excel.load_table(filename, head_row=i)[sheetname]
        table = list(filter(lambda it: it['ФИО'] != None, table))
        table = list(filter(lambda it: it['Тип диплома'] != None, table))
        peoples.extend(list(map(lambda it: sheetname + ' ' + it['ФИО'], table)))

    return peoples


def make_diplom(filename, target, path_template, dir_out, filename_out, reg_number, number, subject):
    wb = load_workbook(filename=filename, read_only=True, data_only=True)

    peoples = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        rows = tuple(ws.rows)

        i = find_head(rows)

        table = load_excel.load_table(filename, head_row=i)[sheetname]
        table = list(filter(lambda it: it['ФИО'] != None, table))
        table = list(filter(lambda it: it['Тип диплома'] != None, table))
        for k in range(len(table)):
            table[k]['Имя листа'] = sheetname
        peoples.extend(table)

    target = target.split("\n")
    target = list(map(lambda it: it.strip(), target))
    target = list(filter(lambda it: it != "", target))
    peoples = list(filter(lambda it: (it['Имя листа'] + " " + it['ФИО']) in target, peoples))

    data = load_excel.load_table(filename=PATH_TO_DB_SCHOOL_SUBJECTS)['Данные']
    data = list(filter(lambda item: item['Дата награждения'] != None, data))
    data = list(filter(lambda item: item['Название предмета д.п.'].lower() == subject.lower(), data))
    if len(data) != 1: raise Exception(11)
    sub = data[0]
    name = sub['Название предмета д.п.']
    name = name[0:1].lower() + name[1:]
    out = ""
    template_document_win = ''
    template_document_no_win = ''

    preview_win = []
    preview_no_win = []

    date = sub['Дата награждения']
    reg_number = int(reg_number)
    number = int(number)
    for people in peoples:
        if people['ФИО'] == 'Бит - Зая Георгий Александрович':
            fio = []
            fio.append('Бит - Зая')
            fio.append('Георгий')
            fio.append('Александрович')
        else:
            fio = people['ФИО'].split(' ')
            fio = list(filter(lambda it: it != "", fio))
            while (len(fio) < 3): fio.append('')

        school = people['Учебное заведение (полностью)']
        school = school.split("\n", 3)
        school = list(map(lambda it: it.strip(), school))
        school = list(map(lambda it: " ".join(it.split("\n")), school))
        school = list(filter(lambda it: it != "", school))
        school.append(people['Муниципалитет'])
        while (len(school) < 5): school.append("")

        num = "24 %s РЭ %06d" % (people['Тип диплома'][:2].upper(), number)
        diplom = "\Diplom" + "{%s}" * 16
        diplom = diplom % (
            name,
            fio[0],
            fio[1],
            " ".join(fio[2:]),
            people['Класс'],
            school[0],
            school[1],
            school[2],
            school[3],
            school[4],
            str(date.day),
            date_to_rus_string.month_to_rus_string(date),
            str(date.year % 100),
            sub['Место проведения'],
            num,
            str(reg_number)
        )
        preview_temp = "%s %s (%s класс, %s, %s)" % (
            fio[0], fio[1], people['Класс'], people['Учебное заведение'], people['Муниципалитет'])
        if people['Тип диплома'].lower() == "победитель":
            template_document_win += diplom + '\n'
            preview_win.append(preview_temp)
        else:
            template_document_no_win += diplom + '\n'
            preview_no_win.append(preview_temp)

        out_people = str(reg_number)
        out_people += '\t' + num
        out_people += '\t' + name
        out_people += '\t' + people['ФИО']
        out_people += '\t' + people['Учебное заведение'] + ", " + people['Муниципалитет'] + ", " + str(
            people['Класс']) + " кл."
        out += out_people + '\n'
        number += 1
        reg_number += 1

    make_pdf.make_pdf(path_template, dir_out, filename_out + " (победители)",
                      {"%%{{template_document}}": template_document_win})
    path = make_pdf.make_pdf(path_template, dir_out, filename_out + " (призёры)",
                             {"%%{{template_document}}": template_document_no_win})

    tt = sub['Название предмета д.п.']
    tt = tt[0].lower() + tt[1:]
    preview = ''
    preview += 'Победителями олимпиады по %s стали: %s' % (tt, ', '.join(preview_win))
    preview += '\n\n\n'
    preview += 'Дипломы призёров получили следующие участники: %s' % (', '.join(preview_no_win))
    return (out, path, preview)
